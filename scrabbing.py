from finviz import FinViz
import argparse
from threading import Thread
import time
from openpyxl import Workbook, load_workbook
import os.path


def to_excel(filename, sheetname, values, verbose=False):
    assert not (filename is None or filename == ''), 'filename is empty'
    assert not (sheetname is None or sheetname == ''), 'sheetname is empty'

    if len(values) == 0:
        return

    if verbose:
        print('Export to excel file =', filename, ' sheetname =', sheetname)

    if os.path.exists(filename) and os.path.isfile(filename):
        # add data in exists file
        wb = load_workbook(filename=filename)
        if sheetname in wb.sheetnames:
            del wb[sheetname]
        ws = wb.create_sheet(sheetname)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname

    # write header of table
    ws.cell(row=1, column=1).value = 'Ticker'
    for k1, v1 in values.items():
        for i, k in enumerate(v1.keys()):
            ws.cell(row=1, column=i + 2).value = k
        break

    line = 2
    for k, v1 in values.items():
        ws.cell(row=line, column=1).value = k
        for i, v2 in enumerate(v1.values()):
            ws.cell(row=line, column=i + 2).value = v2
            if type(v2) in ('float', 'int'):
                ws.cell(row=line, column=i + 2).number_format = '0.00'
        line += 1

    wb.save(filename)


def to_stdout(values, space_char):
    if len(values) == 0:
        return

    for k, v in values.items():
        print('Ticker', end=space_char)
        print(k)
        for k2, v2 in v.items():
            print(k2, end=space_char)
            print(v2)
        print()


def get_args():
    parser = argparse.ArgumentParser(description='FinViz scrabing...')
    parser.add_argument(
        'tickers',
        help='tickers of shares, separated by space',
        nargs='+')
    parser.add_argument(
        '-v',
        '--verbose',
        help='show more information about program works',
        action='store_true')
    parser.add_argument(
        '--silent',
        help='drop all message to stdout',
        action='store_true')
    parser.add_argument(
        '--excel',
        help='export data in excel document',
        nargs=2,
        metavar=(
            'filename',
            'sheetname'))
    return parser.parse_args()

def _scrab(tic, values):
    fz = FinViz(tic=tic)
    values[tic] = fz.get_values()

def main():
    args = get_args()
    values = {}

    tickers = ' '.join(args.tickers).split()

    if not args.silent:
        print(*tickers)

    if not args.silent:
        start_time = time.time()

    threads = []
    for tic in tickers:
        t = Thread(target=_scrab, args=(tic, values))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    # sort values by keys
    values = dict(sorted(values.items()))

    if args.verbose:
        print('values = ', values)

    if args.excel is None:
        to_stdout(values, '\t')
    else:
        if not args.silent:
            print('exporting to', args.excel[0])
        to_excel(args.excel[0], args.excel[1], values, args.verbose)

    if not args.silent:
        print(
            'scabbing completed. It takes {} seconds'.format(
                round(
                    time.time() -
                    start_time,
                    2)))


if __name__ == "__main__":
    main()
